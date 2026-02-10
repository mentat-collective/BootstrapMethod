"""Generate the Bootstrap Method Google Sheet (.xlsx) with 4 tabs.

Tab 1: Propeller Blade Measurements → TAF
Tab 2: Glide & Climb Flight Tests → CD0, e
Tab 3: Bootstrap Data Plate Summary
Tab 4: Performance Calculator (full Bootstrap Method computation)

Upload the resulting .xlsx to Google Sheets.
"""

import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, size=12)
SECTION_FONT = Font(bold=True, size=11, color="2F5496")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALC_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
RESULT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_cell(ws, row, col, value, font=None, fill=None, fmt=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = THIN_BORDER
    return cell


def create_tab1_propeller(wb):
    """Tab 1: Propeller Blade Measurements → TAF"""
    ws = wb.active
    ws.title = "Prop Blade → TAF"

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22

    # Title
    ws.merge_cells("A1:F1")
    style_cell(ws, 1, 1, "Propeller Blade Activity Factor (BAF & TAF)", font=TITLE_FONT, border=False)

    # Instructions
    ws.merge_cells("A2:F2")
    style_cell(ws, 2, 1,
               "Measure blade width at each station using calipers. "
               "Yellow cells = your inputs. Blue cells = computed.",
               border=False)

    # --- Prop specs ---
    r = 4
    style_cell(ws, r, 1, "Propeller Specs", font=SECTION_FONT, border=False)
    r = 5
    style_cell(ws, r, 1, "Blade Radius R (inches):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)  # user enters R here (B5)
    r = 6
    style_cell(ws, r, 1, "Number of Blades BB:", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)  # user enters BB here (B6)
    r = 7
    style_cell(ws, r, 1, "Propeller Model:", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)  # user enters model here (B7)

    # --- Station measurements ---
    r = 9
    style_cell(ws, r, 1, "Station Measurements", font=SECTION_FONT, border=False)
    r = 10
    style_cell(ws, r, 1, "Station (x = r/R)", font=BOLD)
    style_cell(ws, r, 2, "r = x × R (in)", font=BOLD)
    style_cell(ws, r, 3, "Blade Width b(x) (in)", font=BOLD)
    style_cell(ws, r, 4, "f(x) = x³ × b(x) / R", font=BOLD)
    style_cell(ws, r, 5, "Trap. Weight", font=BOLD)
    style_cell(ws, r, 6, "Weighted f(x)", font=BOLD)

    stations = [0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50,
                0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85,
                0.90, 0.95, 1.00]

    for i, x in enumerate(stations):
        row = 11 + i
        # Station x
        style_cell(ws, row, 1, x, fmt="0.00")
        # r = x * R  (formula referencing B5)
        style_cell(ws, row, 2, f"=A{row}*$B$5", fill=CALC_FILL, fmt="0.00")
        # Blade width: user input
        style_cell(ws, row, 3, None, fill=INPUT_FILL)
        # f(x) = x³ * b(x) / R
        style_cell(ws, row, 4, f"=A{row}^3*C{row}/$B$5", fill=CALC_FILL, fmt="0.00000")
        # Trapezoidal weight: 1 for first and last, 2 for middle
        weight = 1 if (i == 0 or i == len(stations) - 1) else 2
        style_cell(ws, row, 5, weight)
        # Weighted f(x)
        style_cell(ws, row, 6, f"=D{row}*E{row}", fill=CALC_FILL, fmt="0.00000")

    # --- Results ---
    result_row = 11 + len(stations) + 1  # row 29
    style_cell(ws, result_row, 1, "Results", font=SECTION_FONT, border=False)

    r = result_row + 1  # row 30
    style_cell(ws, r, 1, "Sum of weighted f(x):", font=BOLD)
    style_cell(ws, r, 2, f"=SUM(F11:F{11+len(stations)-1})", fill=CALC_FILL, fmt="0.0000")

    r += 1  # row 31
    # BAF = (78.125 / R) * sum_weighted_f  (Lowry Eq. 6.56)
    style_cell(ws, r, 1, "BAF (Blade Activity Factor):", font=BOLD)
    style_cell(ws, r, 2, f"=78.125/$B$5*B{r-1}", fill=RESULT_FILL, fmt="0.00")
    style_cell(ws, r, 3, "Eq. 6.56: BAF = (78.125/R) × Σ weighted f(x)", border=False)

    r += 1  # row 32
    # TAF = BB * BAF
    style_cell(ws, r, 1, "TAF (Total Activity Factor):", font=BOLD)
    style_cell(ws, r, 2, f"=$B$6*B{r-1}", fill=RESULT_FILL, fmt="0.00")
    style_cell(ws, r, 3, "Eq. 6.55: TAF = BB × BAF", border=False)

    r += 1  # row 33
    # X = 0.001515 * TAF - 0.0880
    style_cell(ws, r, 1, "X (Power Adj. Factor):", font=BOLD)
    style_cell(ws, r, 2, f"=0.001515*B{r-1}-0.0880", fill=RESULT_FILL, fmt="0.0000")
    style_cell(ws, r, 3, "Eq. 6.57: X = 0.001515 × TAF - 0.0880", border=False)

    r += 2  # row 35
    style_cell(ws, r, 1, "Validation:", font=SECTION_FONT, border=False)
    r += 1
    style_cell(ws, r, 1, "Typical GA BAF range: 70-140", border=False)
    r += 1
    style_cell(ws, r, 1, "R182 example: R=41, BB=2 → BAF=97.94, TAF=195.9", border=False)


def create_tab2_flight_tests(wb):
    """Tab 2: Glide & Climb Flight Tests → CD0, e"""
    ws = wb.create_sheet("Flight Tests → CD0, e")

    # Column widths
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["J"].width = 18

    # Title
    ws.merge_cells("A1:H1")
    style_cell(ws, 1, 1, "Glide & Climb Flight Tests", font=TITLE_FONT, border=False)

    ws.merge_cells("A2:H2")
    style_cell(ws, 2, 1,
               "Yellow = inputs. Blue = computed. Green = results. "
               "Glide tests derive CD0 and e. Climb tests are for validation.",
               border=False)

    # === Aircraft constants ===
    r = 4
    style_cell(ws, r, 1, "Aircraft Constants", font=SECTION_FONT, border=False)
    r = 5
    style_cell(ws, r, 1, "Wing area S (ft²):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 6
    style_cell(ws, r, 1, "Wing span B (ft):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 7
    style_cell(ws, r, 1, "Aspect ratio A:", font=BOLD)
    style_cell(ws, r, 2, "=B6^2/B5", fill=CALC_FILL, fmt="0.000")

    # === Test conditions ===
    r = 9
    style_cell(ws, r, 1, "Test Conditions", font=SECTION_FONT, border=False)
    r = 10
    style_cell(ws, r, 1, "Date:", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 11
    style_cell(ws, r, 1, "Top Pressure Alt (ft):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 12
    style_cell(ws, r, 1, "Bottom Pressure Alt (ft):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 13
    style_cell(ws, r, 1, "ΔH pressure (ft):", font=BOLD)
    style_cell(ws, r, 2, "=B11-B12", fill=CALC_FILL, fmt="0.0")
    r = 14
    style_cell(ws, r, 1, "OAT at midpoint (°F):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 15
    style_cell(ws, r, 1, "Mid pressure alt (ft):", font=BOLD)
    style_cell(ws, r, 2, "=(B11+B12)/2", fill=CALC_FILL, fmt="0.0")

    # Standard temp at mid altitude
    r = 16
    style_cell(ws, r, 1, "Std temp at mid alt (°F):", font=BOLD)
    style_cell(ws, r, 2, "=59-0.003566*B15", fill=CALC_FILL, fmt="0.0")

    # Tapeline correction factor: (OAT + 459.7) / (Tstd + 459.7)
    r = 17
    style_cell(ws, r, 1, "Tapeline correction:", font=BOLD)
    style_cell(ws, r, 2, "=(B14+459.7)/(B16+459.7)", fill=CALC_FILL, fmt="0.0000")

    # ΔH tapeline
    r = 18
    style_cell(ws, r, 1, "ΔH tapeline (ft):", font=BOLD)
    style_cell(ws, r, 2, "=B13*B17", fill=CALC_FILL, fmt="0.0")

    # Sigma at mid altitude
    r = 19
    style_cell(ws, r, 1, "σ (density ratio):", font=BOLD)
    style_cell(ws, r, 2, "=(1-0.003566*B15/518.7)^(1/0.234957)", fill=CALC_FILL, fmt="0.0000")

    # Rho
    r = 20
    style_cell(ws, r, 1, "ρ (slug/ft³):", font=BOLD)
    style_cell(ws, r, 2, "=0.002377*B19", fill=CALC_FILL, fmt="0.000000")

    # Empty weight, fuel, occupants for weight computation
    r = 22
    style_cell(ws, r, 1, "Weight Computation", font=SECTION_FONT, border=False)
    r = 23
    style_cell(ws, r, 1, "Empty weight (lbs):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 24
    style_cell(ws, r, 1, "Pilot + pax (lbs):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)
    r = 25
    style_cell(ws, r, 1, "Baggage (lbs):", font=BOLD)
    style_cell(ws, r, 2, None, fill=INPUT_FILL)

    # === IAS to CAS correction ===
    r = 27
    style_cell(ws, r, 1, "IAS → CAS Correction", font=SECTION_FONT, border=False)
    r = 28
    style_cell(ws, r, 1, "Position error (kt):", font=BOLD)
    style_cell(ws, r, 2, 0, fill=INPUT_FILL)
    style_cell(ws, r, 3, "(Enter correction to add; 0 if KIAS ≈ KCAS)", border=False)

    # === GLIDE TEST DATA ===
    r = 30
    style_cell(ws, r, 1, "GLIDE TEST RUNS", font=SECTION_FONT, border=False)
    style_cell(ws, r, 5, "Prop at low RPM, power idle, trimmed & stabilized", border=False)

    r = 31
    # Columns A-H: core data; I-J: regression basis functions
    headers = ["Run #", "Fuel (gal)", "Gross Wt (lbs)", "KIAS",
               "KCAS", "Δt (sec)", "V_TAS (fps)", "KCAS × Δt",
               "V/Δt", "V⁴"]
    for i, h in enumerate(headers):
        style_cell(ws, r, i + 1, h, font=BOLD)

    # 12 glide test rows
    for run in range(1, 13):
        row = 31 + run
        style_cell(ws, row, 1, run)  # Run #
        style_cell(ws, row, 2, None, fill=INPUT_FILL)  # Fuel gal
        # Gross weight = empty + pax + baggage + fuel*6
        style_cell(ws, row, 3, f"=$B$23+$B$24+$B$25+B{row}*6", fill=CALC_FILL, fmt="0.0")
        style_cell(ws, row, 4, None, fill=INPUT_FILL)  # KIAS
        # KCAS = KIAS + position error
        style_cell(ws, row, 5, f"=D{row}+$B$28", fill=CALC_FILL, fmt="0.0")
        style_cell(ws, row, 6, None, fill=INPUT_FILL)  # delta-t
        # V_TAS in ft/sec = (KCAS / sqrt(sigma)) / 0.5924838
        style_cell(ws, row, 7, f"=IFERROR((E{row}/SQRT($B$19))/0.5924838,\"\")",
                   fill=CALC_FILL, fmt="0.00")
        # KCAS * delta-t (for visual inspection of the curve)
        style_cell(ws, row, 8, f"=IFERROR(E{row}*F{row},\"\")", fill=CALC_FILL, fmt="0.0")
        # V_TAS / delta-t  (y for regression: V/Δt = a·V⁴ + b)
        style_cell(ws, row, 9, f"=IFERROR(G{row}/F{row},\"\")", fill=CALC_FILL, fmt="0.000")
        # V_TAS^4 (x for regression)
        style_cell(ws, row, 10, f"=IFERROR(G{row}^4,\"\")", fill=CALC_FILL, fmt="0.0")

    # === Curve Fit: V/Δt = a·V⁴ + b ===
    # From the drag polar: D = CD0·q·S + W²/(q·S·π·A·e)
    # Rate of sink: ROS = D·V/W
    # Through algebra: V/Δt = [CD0·ρ·S/(2·W·ΔH)]·V⁴ + [2·W/(ρ·S·π·A·e·ΔH)]
    # This is linear in V⁴, so standard SLOPE/INTERCEPT regression applies.
    # From the fit coefficients:
    #   V_bg = (b/a)^(1/4)   — exact best glide speed
    #   CD0 = a · 2·W·ΔH / (ρ·S)
    #   e   = 2·W / (b · ρ·S·π·A·ΔH)

    r = 45
    style_cell(ws, r, 1, "Curve Fit: V/Δt = a·V⁴ + b", font=SECTION_FONT, border=False)

    r = 46
    style_cell(ws, r, 1, "Avg gross weight W (lbs):", font=BOLD)
    style_cell(ws, r, 2, "=AVERAGE(C32:C43)", fill=CALC_FILL, fmt="0.0")
    style_cell(ws, r, 3, "(used for CD0/e extraction)", border=False)

    r = 47
    style_cell(ws, r, 1, "a (slope):", font=BOLD)
    style_cell(ws, r, 2, "=SLOPE(I32:I43,J32:J43)", fill=CALC_FILL, fmt="0.000000000")
    style_cell(ws, r, 3, "a = CD0·ρ·S / (2·W·ΔH)", border=False)

    r = 48
    style_cell(ws, r, 1, "b (intercept):", font=BOLD)
    style_cell(ws, r, 2, "=INTERCEPT(I32:I43,J32:J43)", fill=CALC_FILL, fmt="0.000000")
    style_cell(ws, r, 3, "b = 2·W / (ρ·S·π·A·e·ΔH)", border=False)

    # V_bg from curve fit
    r = 49
    style_cell(ws, r, 1, "V_bg TAS (fps):", font=BOLD)
    style_cell(ws, r, 2, "=(B48/B47)^0.25", fill=CALC_FILL, fmt="0.00")
    style_cell(ws, r, 3, "V_bg = (b/a)^(1/4)", border=False)

    r = 50
    style_cell(ws, r, 1, "Vbg (KCAS):", font=BOLD)
    style_cell(ws, r, 2, "=B49*SQRT($B$19)*0.5924838", fill=RESULT_FILL, fmt="0.0")
    style_cell(ws, r, 3, "V_bg_TAS × √σ × 0.5924838", border=False)

    # === CD0 and e from curve fit ===
    r = 52
    style_cell(ws, r, 1, "CD0 and e (from curve fit)", font=SECTION_FONT, border=False)

    # CD0 = a × 2·W·ΔH / (ρ·S)
    r = 53
    style_cell(ws, r, 1, "CD0:", font=BOLD)
    style_cell(ws, r, 2, "=B47*2*B46*$B$18/($B$20*$B$5)", fill=RESULT_FILL, fmt="0.00000")
    style_cell(ws, r, 3, "a × 2·W·ΔH / (ρ·S)", border=False)

    # e = 2·W / (b × ρ·S·π·A·ΔH)
    r = 54
    style_cell(ws, r, 1, "e (efficiency factor):", font=BOLD)
    style_cell(ws, r, 2, "=2*B46/(B48*$B$20*$B$5*PI()*$B$7*$B$18)",
               fill=RESULT_FILL, fmt="0.000")
    style_cell(ws, r, 3, "2·W / (b·ρ·S·π·A·ΔH)", border=False)

    # Max L/D for reference
    r = 55
    style_cell(ws, r, 1, "Max L/D:", font=BOLD)
    style_cell(ws, r, 2, "=1/(2*SQRT(B53/(PI()*$B$7*B54)))", fill=CALC_FILL, fmt="0.0")
    style_cell(ws, r, 3, "1 / (2·√(CD0/(π·A·e)))", border=False)

    # R² for fit quality
    r = 56
    style_cell(ws, r, 1, "R² (fit quality):", font=BOLD)
    style_cell(ws, r, 2, "=RSQ(I32:I43,J32:J43)", fill=CALC_FILL, fmt="0.0000")
    style_cell(ws, r, 3, "Should be > 0.99 for good data", border=False)

    # KCAS×Δt at Vbg (for sanity check vs raw data)
    r = 57
    style_cell(ws, r, 1, "Max KCAS×Δt (raw data):", font=BOLD)
    style_cell(ws, r, 2, "=MAX(H32:H43)", fill=CALC_FILL, fmt="0.0")
    style_cell(ws, r, 3, "Sanity check: Vbg should be near the max row", border=False)

    # === CLIMB TEST DATA (validation) ===
    r = 62
    style_cell(ws, r, 1, "CLIMB TEST RUNS (Validation)", font=SECTION_FONT, border=False)
    style_cell(ws, r, 5, "Full power at 2500 RPM, trimmed & stabilized", border=False)

    r = 63
    climb_headers = ["Run #", "Fuel (gal)", "Gross Wt (lbs)", "KIAS",
                     "KCAS", "Δt (sec)", "ROC (fpm)", "RPM", "% Power"]
    for i, h in enumerate(climb_headers):
        style_cell(ws, r, i + 1, h, font=BOLD)

    # 12 climb test rows
    for run in range(1, 13):
        row = 63 + run
        style_cell(ws, row, 1, run)
        style_cell(ws, row, 2, None, fill=INPUT_FILL)  # Fuel gal
        style_cell(ws, row, 3, f"=$B$23+$B$24+$B$25+B{row}*6", fill=CALC_FILL, fmt="0.0")
        style_cell(ws, row, 4, None, fill=INPUT_FILL)  # KIAS
        style_cell(ws, row, 5, f"=D{row}+$B$28", fill=CALC_FILL, fmt="0.0")
        style_cell(ws, row, 6, None, fill=INPUT_FILL)  # delta-t
        # ROC = ΔH_tapeline / Δt * 60
        style_cell(ws, row, 7, f"=$B$18/F{row}*60", fill=CALC_FILL, fmt="0.0")
        style_cell(ws, row, 8, None, fill=INPUT_FILL)  # RPM
        style_cell(ws, row, 9, None, fill=INPUT_FILL)  # % Power from Dynon

    # === Climb test analysis: derive Vx from measured data ===
    r = 77
    style_cell(ws, r, 1, "Climb Test Analysis", font=SECTION_FONT, border=False)
    r = 78
    style_cell(ws, r, 1, "Best ROC speed (Vy):", font=BOLD)
    # Find KCAS of the row with max ROC. Use INDEX/MATCH.
    style_cell(ws, r, 2,
               "=IFERROR(INDEX(E64:E75,MATCH(MAX(G64:G75),G64:G75,0)),\"\")",
               fill=RESULT_FILL, fmt="0.0")
    style_cell(ws, r, 3, "KCAS", font=BOLD)
    style_cell(ws, r, 4, "KCAS at max ROC", border=False)

    r = 79
    style_cell(ws, r, 1, "Max ROC:", font=BOLD)
    style_cell(ws, r, 2, "=IFERROR(MAX(G64:G75),\"\")", fill=RESULT_FILL, fmt="0.0")
    style_cell(ws, r, 3, "ft/min", font=BOLD)

    r = 80
    style_cell(ws, r, 1, "Best climb angle speed (Vx):", font=BOLD)
    # Climb angle ≈ arcsin(ROC / (V_TAS × 60)) in degrees
    # But for finding the max, we can compute sin(γ) = ROC/(V*60) for each row.
    # V_TAS(fps) = (KCAS / √σ) / 0.5924838
    # Climb angle = DEGREES(ASIN(ROC / (V_TAS * 60)))
    # We need a helper column. Let's put climb angle in column J.
    style_cell(ws, r, 2,
               "=IFERROR(INDEX(E64:E75,MATCH(MAX(J64:J75),J64:J75,0)),\"\")",
               fill=RESULT_FILL, fmt="0.0")
    style_cell(ws, r, 3, "KCAS", font=BOLD)
    style_cell(ws, r, 4, "KCAS at max climb angle", border=False)

    r = 81
    style_cell(ws, r, 1, "Max climb angle:", font=BOLD)
    style_cell(ws, r, 2, "=IFERROR(MAX(J64:J75),\"\")", fill=RESULT_FILL, fmt="0.00")
    style_cell(ws, r, 3, "degrees", font=BOLD)

    # Add climb angle column header and formulas
    style_cell(ws, 63, 10, "Climb Angle (°)", font=BOLD)
    for run in range(1, 13):
        row = 63 + run
        # Climb angle = DEGREES(ATAN(ROC / (V_TAS * 60)))
        # V_TAS = (KCAS / sqrt(sigma)) / 0.5924838
        style_cell(ws, row, 10,
                   f"=IFERROR(DEGREES(ATAN(G{row}/(((E{row}/SQRT($B$19))/0.5924838)*60))),\"\")",
                   fill=CALC_FILL, fmt="0.00")

    r = 83
    style_cell(ws, r, 1, "Compare these against bootstrap predictions:", border=False)
    style_cell(ws, r, 4,
               "Run the Clojure calculator at the same W, h, RPM, % power",
               border=False)

    # === CHARTS ===
    # Chart 1: V/Δt vs V⁴ (the regression basis — shows linearity)
    chart1 = ScatterChart()
    chart1.title = "Glide Regression: V/Δt vs V⁴"
    chart1.x_axis.title = "V⁴ (fps⁴)"
    chart1.y_axis.title = "V/Δt (fps/sec)"
    chart1.width = 18
    chart1.height = 12

    x_data = Reference(ws, min_col=10, min_row=32, max_row=43)  # V⁴ (col J)
    y_data = Reference(ws, min_col=9, min_row=32, max_row=43)   # V/Δt (col I)
    series1 = Series(y_data, x_data, title="Glide data")
    series1.graphicalProperties.line.noFill = True  # scatter, no line between points
    tl = Trendline(trendlineType="linear", dispRSqr=True, dispEq=True)
    tl.graphicalProperties = GraphicalProperties()
    tl.graphicalProperties.line.solidFill = "FF0000"  # red trendline
    series1.trendline = tl
    chart1.series.append(series1)
    ws.add_chart(chart1, "A86")

    # Chart 2: KCAS × Δt vs KCAS (intuitive view — peak = Vbg)
    chart2 = ScatterChart()
    chart2.title = "Glide Endurance: KCAS × Δt vs KCAS"
    chart2.x_axis.title = "KCAS"
    chart2.y_axis.title = "KCAS × Δt"
    chart2.width = 18
    chart2.height = 12

    x_data2 = Reference(ws, min_col=5, min_row=32, max_row=43)  # KCAS (col E)
    y_data2 = Reference(ws, min_col=8, min_row=32, max_row=43)  # KCAS×Δt (col H)
    series2 = Series(y_data2, x_data2, title="Glide data")
    chart2.series.append(series2)
    ws.add_chart(chart2, "A102")

    # Chart 3: Climb ROC vs KCAS
    chart3 = ScatterChart()
    chart3.title = "Climb Rate vs Airspeed"
    chart3.x_axis.title = "KCAS"
    chart3.y_axis.title = "Rate of Climb (ft/min)"
    chart3.width = 18
    chart3.height = 12

    x_data3 = Reference(ws, min_col=5, min_row=64, max_row=75)  # KCAS (col E)
    y_data3 = Reference(ws, min_col=7, min_row=64, max_row=75)  # ROC (col G)
    series3 = Series(y_data3, x_data3, title="Climb data")
    chart3.series.append(series3)
    ws.add_chart(chart3, "A118")


def create_tab3_data_plate(wb):
    """Tab 3: Bootstrap Data Plate Summary"""
    ws = wb.create_sheet("Data Plate")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40

    # Title
    ws.merge_cells("A1:D1")
    style_cell(ws, 1, 1, "Bootstrap Data Plate", font=TITLE_FONT, border=False)

    ws.merge_cells("A2:D2")
    style_cell(ws, 2, 1,
               "Yellow = manual inputs. Green = computed from other tabs. "
               "Copy the Clojure map below into your code.",
               border=False)

    # Header
    r = 4
    style_cell(ws, r, 1, "Parameter", font=BOLD)
    style_cell(ws, r, 2, "Value", font=BOLD)
    style_cell(ws, r, 3, "Units", font=BOLD)
    style_cell(ws, r, 4, "Source", font=BOLD)

    # Data rows
    params = [
        ("S (wing area)",         174.0,   "ft²",       "POH / plans",         INPUT_FILL),
        ("B (wing span)",         36.0,    "ft",        "POH / plans",         INPUT_FILL),
        ("P0 (rated power)",      235.0,   "hp",        "POH",                 INPUT_FILL),
        ("N0 (rated RPM)",        2400,    "RPM",       "POH",                 INPUT_FILL),
        ("d (prop diameter)",     6.83,    "ft",        "Measurement",         INPUT_FILL),
        ("CD0",                   "='Flight Tests → CD0, e'!B53", "",  "Tab 2 (curve fit)",  RESULT_FILL),
        ("e",                     "='Flight Tests → CD0, e'!B54", "",  "Tab 2 (curve fit)",  RESULT_FILL),
        ("TAF",                   "='Prop Blade → TAF'!B32",      "",  "Tab 1 (prop measurement)", RESULT_FILL),
        ("Z (fuselage dia / prop dia)", 0.688, "",      "Measurement: fuse_dia / prop_dia", INPUT_FILL),
        ("Tractor? (1=yes, 0=no)", 1,      "",          "Configuration",       INPUT_FILL),
        ("BB (num blades)",       2,       "",           "Observation",         INPUT_FILL),
        ("C (power dropoff)",     0.12,    "",           "Typical normally-aspirated: 0.12", INPUT_FILL),
    ]

    for i, (name, val, units, source, fill) in enumerate(params):
        row = 5 + i
        style_cell(ws, row, 1, name, font=BOLD)
        style_cell(ws, row, 2, val, fill=fill, fmt="0.00000" if name in ("CD0", "e") else "0.00")
        style_cell(ws, row, 3, units)
        style_cell(ws, row, 4, source)

    # Derived: X from TAF
    r = 5 + len(params)
    style_cell(ws, r, 1, "X (power adj. factor)", font=BOLD)
    style_cell(ws, r, 2, "=0.001515*B12-0.0880", fill=RESULT_FILL, fmt="0.0000")
    style_cell(ws, r, 3, "")
    style_cell(ws, r, 4, "Eq 6.57: X = 0.001515 × TAF - 0.0880")

    # Aspect ratio
    r += 1
    style_cell(ws, r, 1, "A (aspect ratio)", font=BOLD)
    style_cell(ws, r, 2, "=B6^2/B5", fill=CALC_FILL, fmt="0.000")
    style_cell(ws, r, 3, "")
    style_cell(ws, r, 4, "B² / S")

    # === Clojure data plate literal ===
    r += 2
    style_cell(ws, r, 1, "Clojure Data Plate (copy-paste):", font=SECTION_FONT, border=False)
    r += 1
    # Each row of the Clojure map
    clj_lines = [
        ('{:S',        '=TEXT(B5,"0.0")',    '  ;; wing area, ft²'),
        (' :B',        '=TEXT(B6,"0.0")',    '  ;; wing span, ft'),
        (' :P0',       '=TEXT(B7,"0.0")',    '  ;; rated MSL power, hp'),
        (' :N0',       '=TEXT(B8,"0")',      '  ;; rated RPM'),
        (' :d',        '=TEXT(B9,"0.000")',  '  ;; prop diameter, ft'),
        (' :CD0',      '=TEXT(B10,"0.00000")', '  ;; parasite drag coefficient'),
        (' :e',        '=TEXT(B11,"0.000")', '  ;; airplane efficiency factor'),
        (' :TAF',      '=TEXT(B12,"0.0")',   '  ;; total activity factor'),
        (' :Z',        '=TEXT(B13,"0.000")', '  ;; fuselage dia / prop dia'),
        (' :tractor?', '=IF(B14=1,"true","false")', ''),
        (' :BB',       '=TEXT(B15,"0")',     '  ;; number of blades'),
        (' :C',        '=TEXT(B16,"0.00")',  '}  ;; altitude power dropoff'),
    ]
    for key, formula, comment in clj_lines:
        ws.cell(row=r, column=1, value=key).font = Font(name="Courier New")
        ws.cell(row=r, column=2, value=formula).font = Font(name="Courier New")
        ws.cell(row=r, column=3, value=comment).font = Font(name="Courier New")
        r += 1


def create_tab4_performance(wb):
    """Tab 4: Performance Calculator — full Bootstrap Method computation.

    Takes a data plate (9 aircraft params + config) and operational variables
    (W, h, N, %power), then computes a full performance table via the GAGPC
    propeller efficiency model. Pre-filled with R182 validation data.
    """
    ws = wb.create_sheet("Performance Calculator")

    # ----- Column widths -----
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    for col_letter in "CDEFGHIJKLMNOPQ":
        ws.column_dimensions[col_letter].width = 13

    # Performance table geometry
    TBL_HDR = 94       # header row
    TBL_START = 95     # first data row (KCAS = 60.0)
    TBL_END = 255      # last data row  (KCAS = 140.0)
    # 161 rows total: (140 - 60) / 0.5 + 1

    # =====================================================================
    # Section 1: GAGPC Polynomial Coefficients (rows 1-12)
    # =====================================================================
    ws.merge_cells("A1:Q1")
    style_cell(ws, 1, 1, "Performance Calculator \u2014 Bootstrap Method",
               font=TITLE_FONT, border=False)
    ws.merge_cells("A2:Q2")
    style_cell(ws, 2, 1,
               "Yellow = inputs. Blue = computed. Green = results. "
               "Change data plate and operational variables to recompute.",
               border=False)

    style_cell(ws, 4, 1, "GAGPC Polynomial Coefficients",
               font=SECTION_FONT, border=False)
    style_cell(ws, 4, 6, "(8 columns \u00d7 7 coefficients \u2014 "
               "Boeing/Uddenberg propeller data)", border=False)

    # CPX breakpoints in row 5, columns C-J
    cpx_breakpoints = [0.15, 0.25, 0.40, 0.60, 0.80, 1.00, 1.20, 1.40]
    style_cell(ws, 5, 1, "CPX \u2192", font=BOLD)
    for i, cpx in enumerate(cpx_breakpoints):
        style_cell(ws, 5, 3 + i, cpx, fill=CALC_FILL, fmt="0.00")

    # GAGPC coefficients: 8 columns (one per CPX) x 7 rows (c0-c6).
    # gagpc[col_idx] = [c0, c1, ..., c6] for that CPX breakpoint.
    gagpc = [
        [-0.027280541925,  1.157818942224, -0.548923123013,  0.038551650269,  0.064580555280, -0.026301243311,  0.003017419881],
        [-0.038502996895,  1.046135815788, -0.485313329667,  0.130100232509, -0.027326610124,  0.003139013961, -0.000131566345],
        [-0.026741905000,  0.717582413500, -0.084673350000, -0.074451680000,  0.026437051000, -0.003537565000,  0.000177733000],
        [ 0.038100252152,  0.199522455590,  0.458898025445, -0.318992736679,  0.081357821405, -0.009083801712,  0.000350613126],
        [ 0.251897476000, -0.561665840000,  1.139055130000, -0.602193330000,  0.144341736000, -0.016194730000,  0.000664038000],
        [ 0.140144145675,  0.071636129794, -0.057356417627,  0.276378945894, -0.159843307011,  0.034239846258, -0.002573002786],
        [-0.705604050000,  2.868232531000, -3.651371295000,  2.487262933000, -0.863421200000,  0.146478331500, -0.009686855000],
        [-2.340532183939,  7.563937897150, -9.020964992475,  5.550045133294, -1.793776917489,  0.291020480454, -0.018743169313],
    ]
    coeff_labels = [
        ("c0", "(constant)"), ("c1", "(h)"), ("c2", "(h\u00b2)"),
        ("c3", "(h\u00b3)"), ("c4", "(h\u2074)"),
        ("c5", "(h\u2075)"), ("c6", "(h\u2076)"),
    ]
    for ci in range(7):
        row = 6 + ci
        style_cell(ws, row, 1, coeff_labels[ci][0], font=BOLD)
        style_cell(ws, row, 2, coeff_labels[ci][1])
        for col_idx in range(8):
            style_cell(ws, row, 3 + col_idx, gagpc[col_idx][ci],
                       fill=CALC_FILL, fmt="0.000000000000")

    # =====================================================================
    # Section 2: SDF Coefficients (rows 14-16)
    # =====================================================================
    style_cell(ws, 14, 1, "Slow-Down Factor (SDF) Coefficients",
               font=SECTION_FONT, border=False)
    style_cell(ws, 14, 6, "Cubic polynomial in Z (Lowry Eq. 6.58/6.59)",
               border=False)

    style_cell(ws, 15, 1, "Tractor:", font=BOLD)
    for i, c in enumerate([1.05263, -0.00722, -0.16462, -0.18341]):
        style_cell(ws, 15, 2 + i, c, fill=CALC_FILL, fmt="0.00000")

    style_cell(ws, 16, 1, "Pusher:", font=BOLD)
    for i, c in enumerate([1.05263, -0.04185, -0.01481, -0.62001]):
        style_cell(ws, 16, 2 + i, c, fill=CALC_FILL, fmt="0.00000")

    # =====================================================================
    # Section 3: Aircraft Data Plate + Operational Variables (rows 18-36)
    #
    # Pre-filled with R182 validation data (Cessna R182 N4697K from
    # bootstp2.xls). Replace with your aircraft's values.
    # =====================================================================
    style_cell(ws, 18, 1, "Aircraft Data Plate",
               font=SECTION_FONT, border=False)
    style_cell(ws, 18, 4, "(linked from Data Plate tab)",
               border=False)

    # Data plate: (row, formula, label, units, fmt)
    # Each cell references the Data Plate tab directly.
    # Row mapping:  S=B19, B=B20, P0=B21, N0=B22, d=B23,
    #               CD0=B24, e=B25, TAF=B26, Z=B27,
    #               Tractor?=B28, BB=B29, C=B30
    dp_params = [
        (19, "='Data Plate'!B5",  "S (wing area)",          "ft\u00b2", "0.0"),
        (20, "='Data Plate'!B6",  "B (wing span)",           "ft",       "0.0"),
        (21, "='Data Plate'!B7",  "P0 (rated power)",        "hp",       "0.0"),
        (22, "='Data Plate'!B8",  "N0 (rated RPM)",          "RPM",      "0"),
        (23, "='Data Plate'!B9",  "d (prop diameter)",        "ft",       "0.000"),
        (24, "='Data Plate'!B10", "CD0",                      "",         "0.00000"),
        (25, "='Data Plate'!B11", "e",                         "",         "0.000"),
        (26, "='Data Plate'!B12", "TAF",                       "",         "0.0"),
        (27, "='Data Plate'!B13", "Z (fuse dia / prop dia)",   "",         "0.000"),
        (28, "='Data Plate'!B14", "Tractor? (1=yes, 0=no)",    "",         "0"),
        (29, "='Data Plate'!B15", "BB (num blades)",            "",         "0"),
        (30, "='Data Plate'!B16", "C (power dropoff)",          "",         "0.00"),
    ]
    for row, formula, label, units, fmt in dp_params:
        style_cell(ws, row, 1, label, font=BOLD)
        style_cell(ws, row, 2, formula, fill=RESULT_FILL, fmt=fmt)
        if units:
            style_cell(ws, row, 3, units, border=False)

    # Operational variables
    style_cell(ws, 32, 1, "Operational Variables",
               font=SECTION_FONT, border=False)
    style_cell(ws, 32, 4, "(change these to explore different conditions)",
               border=False)

    # W=B33, h=B34, N=B35, %Power=B36
    ops = [
        (33, 3100.0, "W (gross weight)",  "lbs", "0.0"),
        (34, 8000,   "h (density alt)",   "ft",  "0"),
        (35, 2300,   "N (RPM)",           "",     "0"),
        (36, 0.65,   "% Power (0\u20131)", "",    "0.00"),
    ]
    for row, default, label, units, fmt in ops:
        style_cell(ws, row, 1, label, font=BOLD)
        style_cell(ws, row, 2, default, fill=INPUT_FILL, fmt=fmt)
        if units:
            style_cell(ws, row, 3, units, border=False)

    # =====================================================================
    # Section 4: Computed Constants (rows 38-52)
    # =====================================================================
    style_cell(ws, 38, 1, "Computed Constants",
               font=SECTION_FONT, border=False)

    cc = [
        (39, "\u03c3 (density ratio)",
         "=(1-0.003566*B34/518.7)^(1/0.234957)",
         "0.0000", "(1 - lapse\u00b7h/T\u2080)^(1/0.235)"),
        (40, "\u03c1 (slug/ft\u00b3)",
         "=0.002377*B39",
         "0.000000", ""),
        (41, "\u03c6 (power lapse)",
         "=(B39-B30)/(1-B30)",
         "0.0000", "Gagg-Ferrar: (\u03c3 - C)/(1 - C)"),
        (42, "X (power adj. factor)",
         "=0.001515*B26-0.0880",
         "0.0000", "Eq 6.57"),
        (43, "A (aspect ratio)",
         "=B20^2/B19",
         "0.000", "B\u00b2 / S"),
        (44, "SDF",
         "=IF(B28=1,"
         "B15+C15*B27+D15*B27^2+E15*B27^3,"
         "B16+C16*B27+D16*B27^2+E16*B27^3)",
         "0.000", "Eq 6.58/6.59"),
        (45, "P (ft\u00b7lbf/s)",
         "=B36*B21*550",
         "0.0", "%power \u00d7 P\u2080 \u00d7 550"),
        (46, "n (rev/s)",
         "=B35/60",
         "0.000", "N / 60"),
        (47, "CP (power coeff.)",
         "=B45/(B40*B46^3*B23^5)",
         "0.00000", "P / (\u03c1\u00b7n\u00b3\u00b7d\u2075)"),
        (48, "CPX",
         "=MIN(MAX(B47/B42,0.15),1.40)",
         "0.0000", "CP / X (clamped to GAGPC range)"),
        (49, "GAGPC bracket idx",
         "=MIN(MATCH(B48,$C$5:$J$5,1),7)",
         "0", "MATCH position in CPX breakpoints"),
        (50, "CPX lower",
         "=INDEX($C$5:$J$5,1,B49)",
         "0.00", ""),
        (51, "CPX upper",
         "=INDEX($C$5:$J$5,1,B49+1)",
         "0.00", ""),
        (52, "Interp. fraction",
         "=(B48-B50)/(B51-B50)",
         "0.0000", "(CPX - CPX_lo) / (CPX_hi - CPX_lo)"),
    ]
    for row, label, formula, fmt, note in cc:
        style_cell(ws, row, 1, label, font=BOLD)
        style_cell(ws, row, 2, formula, fill=CALC_FILL, fmt=fmt)
        if note:
            style_cell(ws, row, 3, note, border=False)

    # =====================================================================
    # Section 5: Optimum V-Speeds (rows 54-59)
    #
    # Performance table columns: A=KCAS, N=ROC, O=AOC, P=ROS, Q=AOG
    # =====================================================================
    style_cell(ws, 54, 1, "Optimum V-Speeds",
               font=SECTION_FONT, border=False)

    ts, te = TBL_START, TBL_END
    vspeeds = [
        # (row, label, kcas_formula, unit1, val_formula, unit2)
        (55, "Vy (best ROC):",
         f"=INDEX(A{ts}:A{te},MATCH(MAX(N{ts}:N{te}),N{ts}:N{te},0))",
         "KCAS", f"=MAX(N{ts}:N{te})", "ft/min"),
        (56, "Vx (best AOC):",
         f"=INDEX(A{ts}:A{te},MATCH(MAX(O{ts}:O{te}),O{ts}:O{te},0))",
         "KCAS", f"=MAX(O{ts}:O{te})", "degrees"),
        (57, "Vbg (best glide):",
         f"=INDEX(A{ts}:A{te},MATCH(MIN(Q{ts}:Q{te}),Q{ts}:Q{te},0))",
         "KCAS", f"=MIN(Q{ts}:Q{te})", "degrees"),
        (58, "Vmd (min sink):",
         f"=INDEX(A{ts}:A{te},MATCH(MIN(P{ts}:P{te}),P{ts}:P{te},0))",
         "KCAS", f"=MIN(P{ts}:P{te})", "ft/min"),
        (59, "VM (max level):",
         f'=MAXIFS(A{ts}:A{te},N{ts}:N{te},">0")',
         "KCAS", None, None),
    ]
    for row, label, kcas_f, u1, val_f, u2 in vspeeds:
        style_cell(ws, row, 1, label, font=BOLD)
        style_cell(ws, row, 2, kcas_f, fill=RESULT_FILL, fmt="0.0")
        style_cell(ws, row, 3, u1, font=BOLD)
        if val_f:
            style_cell(ws, row, 4, val_f, fill=RESULT_FILL,
                       fmt="0.0" if "ft/min" == u2 else "0.00")
            style_cell(ws, row, 5, u2)

    # =====================================================================
    # Section 6: Three-Way Validation (rows 61-91)
    #
    # Expected values from bootstp2.xls AND Clojure performance_test.clj.
    # =====================================================================
    style_cell(ws, 61, 1,
               "Validation: Spreadsheet vs Clojure vs bootstp2.xls",
               font=SECTION_FONT, border=False)
    style_cell(ws, 62, 1,
               "R182 at W=3100, h=8000, N=2300, 65% power",
               border=False)

    # Column headers
    for ci, hdr in enumerate(["Item", "Expected", "Computed", "Delta"], 1):
        style_cell(ws, 63, ci, hdr, font=BOLD)

    def val_row(row, name, expected, computed, fmt):
        style_cell(ws, row, 1, name)
        style_cell(ws, row, 2, expected, fmt=fmt)
        style_cell(ws, row, 3, computed, fill=CALC_FILL, fmt=fmt)
        style_cell(ws, row, 4, f"=ABS(C{row}-B{row})", fill=CALC_FILL, fmt=fmt)

    # Part A: Constants
    val_row(64, "\u03c3",    0.786,    "=B39", "0.0000")
    val_row(65, "\u03c1",    0.001868, "=B40", "0.000000")
    val_row(66, "\u03c6",    0.7568,   "=B41", "0.0000")
    val_row(67, "X",         0.2088,   "=B42", "0.0000")
    val_row(68, "SDF",       0.910,    "=B44", "0.000")

    # Part B: Performance at 60 KCAS
    style_cell(ws, 70, 1, "Performance at 60 KCAS",
               font=SECTION_FONT, border=False)
    style_cell(ws, 70, 3, "(bootstp2.xls row 101)", border=False)

    # Column map: H=eta, I=Thrust, K=Dp, L=Di, M=Drag, N=ROC, P=ROS, Q=AOG
    def idx60(col):
        return f"=INDEX({col}{ts}:{col}{te},MATCH(60,A{ts}:A{te},0))"

    val_row(71, "\u03b7 (eta)",  0.617,   idx60("H"), "0.000")
    val_row(72, "Thrust",        453.50,  idx60("I"), "0.00")
    val_row(73, "Dp",            60.95,   idx60("K"), "0.00")
    val_row(74, "Di",            268.96,  idx60("L"), "0.00")
    val_row(75, "Drag",          329.91,  idx60("M"), "0.00")
    val_row(76, "ROC",           273.23,  idx60("N"), "0.0")
    val_row(77, "ROS",           729.37,  idx60("P"), "0.0")
    val_row(78, "AOG",           6.109,   idx60("Q"), "0.000")

    # Part C: V-Speed comparison
    style_cell(ws, 80, 1, "V-Speed Comparison",
               font=SECTION_FONT, border=False)
    val_row(81, "Vy KCAS",  77.0,  "=B55", "0.0")
    val_row(82, "Vy ROC",   371.7, "=D55", "0.0")
    val_row(83, "Vx KCAS",  69.5,  "=B56", "0.0")
    val_row(84, "Vx AOC",   2.55,  "=D56", "0.00")
    val_row(85, "Vbg KCAS", 87.0,  "=B57", "0.0")
    val_row(86, "Vbg AOG",  4.74,  "=D57", "0.00")
    val_row(87, "Vmd KCAS", 66.0,  "=B58", "0.0")
    val_row(88, "Vmd ROS",  719.9, "=D58", "0.0")
    val_row(89, "VM KCAS",  111.5, "=B59", "0.0")

    style_cell(ws, 91, 1,
               "Deltas < 0.5 for KCAS (0.5 kt resolution), < 1% for others",
               border=False)

    # =====================================================================
    # Section 7: Performance Table (rows 93-255)
    #
    # Columns: A=KCAS, B=KTAS, C=V_fps, D=J, E=h, F=eta_lo, G=eta_hi,
    #          H=eta, I=Thrust, J=q, K=Dp, L=Di, M=Drag,
    #          N=ROC, O=AOC, P=ROS, Q=AOG
    # =====================================================================
    style_cell(ws, 93, 1, "Performance Table",
               font=SECTION_FONT, border=False)
    style_cell(ws, 93, 5,
               "KCAS 60\u2013140 (step 0.5) at current conditions",
               border=False)

    headers = [
        "KCAS", "KTAS", "V_fps", "J", "h",
        "\u03b7_lo", "\u03b7_hi", "\u03b7",
        "Thrust", "q", "Dp", "Di", "Drag",
        "ROC", "AOC", "ROS", "AOG",
    ]
    for i, hdr in enumerate(headers):
        style_cell(ws, TBL_HDR, 1 + i, hdr, font=BOLD)

    # Column format map (1-indexed column to number format)
    col_fmts = {
        1: "0.0",      # KCAS
        2: "0.0",      # KTAS
        3: "0.00",     # V_fps
        4: "0.000",    # J
        5: "0.000",    # h
        6: "0.0000",   # eta_lo
        7: "0.0000",   # eta_hi
        8: "0.000",    # eta
        9: "0.0",      # Thrust
        10: "0.00",    # q
        11: "0.00",    # Dp
        12: "0.00",    # Di
        13: "0.0",     # Drag
        14: "0.0",     # ROC
        15: "0.00",    # AOC
        16: "0.0",     # ROS
        17: "0.00",    # AOG
    }

    # --- Helper: Horner polynomial formula for GAGPC ---
    def horner(row, idx_offset):
        """Build Horner's method formula for GAGPC polynomial evaluation.

        GAGPC coefficients: $C$6:$J$12 (c0 in row 6, c6 in row 12).
        Bracket index: $B$49 (from MATCH on CPX breakpoints).
        h (speed-power coefficient): column E of the given row.

        Horner form: c0 + h*(c1 + h*(c2 + h*(c3 + h*(c4 + h*(c5 + h*c6)))))
        """
        h = f"E{row}"
        idx = "$B$49" if idx_offset == 0 else "$B$49+1"

        # Build from c6 (row 12) inward to c1 (row 7)
        result = f"INDEX($C$12:$J$12,1,{idx})"  # c6
        for cr in range(11, 6, -1):  # c5 (row 11) down to c1 (row 7)
            result = f"INDEX($C${cr}:$J${cr},1,{idx})+{h}*({result})"
        # Final outer wrap: c0 (row 6) + h*(c1 + ...)
        result = f"INDEX($C$6:$J$6,1,{idx})+{h}*({result})"
        return f"={result}"

    # --- Generate 161 data rows ---
    for i in range(161):
        row = TBL_START + i
        kcas = 60.0 + i * 0.5

        formulas = {
            1: kcas,                                            # KCAS
            2: f"=A{row}/SQRT($B$39)",                         # KTAS
            3: f"=B{row}/0.5924838",                           # V_fps
            4: f"=C{row}/($B$46*$B$23)",                       # J
            5: f"=D{row}/$B$47^(1/3)",                         # h
            6: horner(row, 0),                                  # eta_lo
            7: horner(row, 1),                                  # eta_hi
            8: f"=$B$44*(F{row}+$B$52*(G{row}-F{row}))",       # eta
            9: f"=H{row}*$B$45/C{row}",                        # Thrust
            10: f"=0.5*$B$40*C{row}^2",                        # q
            11: f"=$B$24*J{row}*$B$19",                         # Dp
            12: f"=$B$33^2/(J{row}*$B$19*PI()*$B$43*$B$25)",   # Di
            13: f"=K{row}+L{row}",                              # Drag
            14: f"=(I{row}-M{row})*C{row}/$B$33*60",           # ROC
            15: f"=DEGREES(ASIN(MIN(1,MAX(-1,"                  # AOC
                f"(I{row}-M{row})/$B$33))))",
            16: f"=M{row}*C{row}/$B$33*60",                    # ROS
            17: f"=DEGREES(ASIN(MIN(1,MAX(-1,"                  # AOG
                f"M{row}/$B$33))))",
        }

        for col in range(1, 18):
            val = formulas[col]
            fill = None if col == 1 else CALC_FILL
            style_cell(ws, row, col, val, fill=fill, fmt=col_fmts[col])

    # =====================================================================
    # Section 8: Charts (after the performance table)
    # =====================================================================

    chart_row = TBL_END + 2  # row 257

    # Chart 1: Thrust & Drag vs Airspeed
    chart1 = ScatterChart()
    chart1.title = "Thrust & Drag vs Airspeed"
    chart1.x_axis.title = "KCAS"
    chart1.y_axis.title = "Force (lbf)"
    chart1.width = 20
    chart1.height = 14
    chart1.style = 2

    x_data = Reference(ws, min_col=1, min_row=TBL_START, max_row=TBL_END)
    thrust_ref = Reference(ws, min_col=9, min_row=TBL_START, max_row=TBL_END)
    drag_ref = Reference(ws, min_col=13, min_row=TBL_START, max_row=TBL_END)

    s_thrust = Series(thrust_ref, x_data, title="Thrust")
    s_drag = Series(drag_ref, x_data, title="Drag")
    chart1.series.append(s_thrust)
    chart1.series.append(s_drag)
    ws.add_chart(chart1, f"A{chart_row}")

    # Chart 2: Rate of Climb vs Airspeed
    chart2 = ScatterChart()
    chart2.title = "Rate of Climb vs Airspeed"
    chart2.x_axis.title = "KCAS"
    chart2.y_axis.title = "ROC (ft/min)"
    chart2.width = 20
    chart2.height = 14
    chart2.style = 2

    roc_ref = Reference(ws, min_col=14, min_row=TBL_START, max_row=TBL_END)
    s_roc = Series(roc_ref, x_data, title="ROC")
    chart2.series.append(s_roc)
    ws.add_chart(chart2, f"J{chart_row}")


def create_tab0_instructions(wb):
    """Tab 0: Instructions — overview of the Bootstrap Method and workflow."""
    ws = wb.create_sheet("Instructions", 0)

    ws.column_dimensions["A"].width = 100
    ws.sheet_properties.tabColor = "2F5496"

    r = 1
    style_cell(ws, r, 1,
               "Bootstrap Method — Aircraft Performance Calculator",
               font=TITLE_FONT, border=False)

    r = 3
    style_cell(ws, r, 1, "What is the Bootstrap Method?",
               font=SECTION_FONT, border=False)
    r = 4
    ws.cell(row=r, column=1).value = (
        "The Bootstrap Method is an aircraft performance prediction technique for "
        "constant-speed propeller airplanes, developed by John T. Lowry in "
        "Performance of Light Aircraft (AIAA, 1999). It derives a complete performance "
        "envelope — thrust, drag, climb, glide, and optimum V-speeds — from a small set "
        "of flight-test-derived parameters called the 'bootstrap data plate.'")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 60

    r = 6
    ws.cell(row=r, column=1).value = (
        "IMPORTANT: This spreadsheet implements the constant-speed propeller version "
        "of the Bootstrap Method using the Boeing/Uddenberg GAGPC propeller efficiency "
        "model. It is NOT suitable for fixed-pitch propeller aircraft, which require a "
        "different propeller model (see Lowry Ch. 5 or the AvWeb Part 1 article below).")
    ws.cell(row=r, column=1).font = Font(bold=True, color="CC0000")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 50

    r = 8
    style_cell(ws, r, 1, "Workflow", font=SECTION_FONT, border=False)
    steps = [
        ("Step 1: Measure your propeller (Tab: Prop Blade → TAF)",
         "Use calipers to measure blade width at 17 standard stations along "
         "the blade. The spreadsheet computes the Blade Activity Factor (BAF), "
         "Total Activity Factor (TAF), and the power adjustment factor X. "
         "This only needs to be done once per propeller."),
        ("Step 2: Fly glide and climb tests (Tab: Flight Tests → CD0, e)",
         "Fly a series of timed glides at different airspeeds with prop at low RPM "
         "and power idle. The spreadsheet uses a V/Δt = a·V⁴ + b curve fit to "
         "extract the parasite drag coefficient (CD0) and airplane efficiency factor (e). "
         "Optionally fly climb tests at a known power setting for validation. "
         "Repeat if you make aerodynamic changes (fairings, seals, etc.)."),
        ("Step 3: Review the data plate (Tab: Data Plate)",
         "The Data Plate tab assembles all nine bootstrap parameters from your "
         "measurements and flight tests. Yellow cells are manual inputs (from your "
         "POH or measurements); green cells are computed from other tabs. A copy-paste "
         "Clojure map literal is provided for use with the companion calculator code."),
        ("Step 4: Explore performance (Tab: Performance Calculator)",
         "The Performance Calculator computes a full performance table at any "
         "combination of weight, density altitude, RPM, and percent power. It "
         "finds all five optimum V-speeds (Vy, Vx, Vbg, Vmd, VM) and includes "
         "validation against the R182 reference data from Lowry's book. Data plate "
         "values are linked from the Data Plate tab — change them there."),
    ]

    for title, desc in steps:
        r += 1
        ws.cell(row=r, column=1).value = title
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        r += 1
        ws.cell(row=r, column=1).value = desc
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 55
        r += 1  # blank row

    r += 1
    style_cell(ws, r, 1, "V-Speeds", font=SECTION_FONT, border=False)
    r += 1
    ws.cell(row=r, column=1).value = (
        "The calculator finds five optimum speeds from the performance table:")
    vspeeds = [
        "Vy — Best Rate of Climb: maximum altitude gain per minute. Use for normal climbs.",
        "Vx — Best Angle of Climb: steepest climb gradient. Use for obstacle clearance.",
        "Vbg — Best Glide: maximum L/D. Parasite drag = induced drag at this speed. "
        "Use after engine failure to maximize glide distance.",
        "Vmd — Minimum Descent Rate: minimizes altitude lost per unit time (slower "
        "than Vbg). Use to stay aloft the longest.",
        "VM — Max Level Flight Speed: highest airspeed where thrust ≥ drag.",
    ]
    for v in vspeeds:
        r += 1
        ws.cell(row=r, column=1).value = f"  • {v}"
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 30

    r += 2
    style_cell(ws, r, 1, "References", font=SECTION_FONT, border=False)
    BOOK_URL = ("https://github.com/mentat-collective/BootstrapMethod/"
                "blob/main/PerfOfLightAircraft.pdf")
    AVWEB_1 = ("https://avweb.com/features_old/the-bootstrap-approach-to-aircraft-"
               "performancepart-one-fixed-pitch-propeller-airplanes/")
    AVWEB_2 = ("https://avweb.com/features_old/the-bootstrap-approach-to-aircraft-"
               "performancepart-two-constant-speed-propeller-airplanes/")
    refs = [
        ("John T. Lowry, Performance of Light Aircraft (AIAA, 1999):", None),
        (f"  {BOOK_URL}", BOOK_URL),
        ("AvWeb: \"The Bootstrap Approach — Part 1: "
         "Fixed-Pitch Propeller Airplanes\":", None),
        (f"  {AVWEB_1}", AVWEB_1),
        ("AvWeb: \"The Bootstrap Approach — Part 2: "
         "Constant-Speed Propeller Airplanes\":", None),
        (f"  {AVWEB_2}", AVWEB_2),
    ]
    for text, url in refs:
        r += 1
        cell = ws.cell(row=r, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True)
        if url:
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[r].height = 30

    r += 2
    style_cell(ws, r, 1, "Color Key", font=SECTION_FONT, border=False)
    r += 1
    style_cell(ws, r, 1, "Yellow cells = your inputs (replace with your values)",
               fill=INPUT_FILL)
    r += 1
    style_cell(ws, r, 1, "Blue cells = computed values (do not edit)",
               fill=CALC_FILL)
    r += 1
    style_cell(ws, r, 1, "Green cells = key results or cross-tab links",
               fill=RESULT_FILL)


def main():
    wb = openpyxl.Workbook()
    create_tab1_propeller(wb)
    create_tab2_flight_tests(wb)
    create_tab3_data_plate(wb)
    create_tab4_performance(wb)
    create_tab0_instructions(wb)

    import os
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bootstrap_method.xlsx")
    wb.save(out)
    print(f"Saved: {out}")
    print("Upload this file to Google Sheets.")


if __name__ == "__main__":
    main()
