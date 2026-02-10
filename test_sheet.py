"""Tests for bootstrap_method.xlsx generation.

Verifies cross-sheet links, R182 defaults, and formula integrity.
Run with: uv run python test_sheet.py
"""

import os
import subprocess
import sys
import openpyxl

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(PROJECT_DIR, "bootstrap_method.xlsx")

passed = 0
failed = 0


def check(name, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
    else:
        failed += 1
        msg = f"FAIL: {name}"
        if detail:
            msg += f" — {detail}"
        print(msg)


def main():
    # Regenerate the spreadsheet
    result = subprocess.run(
        ["uv", "run", "python", "create_sheet.py"],
        capture_output=True, text=True,
        cwd=PROJECT_DIR,
    )
    if result.returncode != 0:
        print(f"create_sheet.py failed:\n{result.stderr}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX)

    # ── Tab references ──
    dp = wb["Data Plate"]
    pc = wb["Performance Calculator"]

    # ================================================================
    # 1. Data Plate: R182 defaults present
    # ================================================================
    r182_defaults = {
        "B5": 174.0,   # S
        "B6": 36.0,    # B
        "B7": 235.0,   # P0
        "B8": 2400,    # N0
        "B9": 6.83,    # d
        "B13": 0.688,  # Z
        "B14": 1,      # Tractor?
        "B15": 2,      # BB
        "B16": 0.12,   # C
    }
    for cell, expected in r182_defaults.items():
        val = dp[cell].value
        check(f"Data Plate {cell} = {expected}",
              val == expected,
              f"got {val!r}")

    # Data Plate computed cells reference other tabs
    check("Data Plate B10 (CD0) refs Tab 2",
          isinstance(dp["B10"].value, str) and "Flight Tests" in dp["B10"].value,
          f"got {dp['B10'].value!r}")
    check("Data Plate B11 (e) refs Tab 2",
          isinstance(dp["B11"].value, str) and "Flight Tests" in dp["B11"].value,
          f"got {dp['B11'].value!r}")
    check("Data Plate B12 (TAF) refs Tab 1",
          isinstance(dp["B12"].value, str) and "Prop Blade" in dp["B12"].value,
          f"got {dp['B12'].value!r}")

    # ================================================================
    # 2. Performance Calculator: data plate cells link to Data Plate
    # ================================================================
    dp_links = {
        "B19": "='Data Plate'!B5",
        "B20": "='Data Plate'!B6",
        "B21": "='Data Plate'!B7",
        "B22": "='Data Plate'!B8",
        "B23": "='Data Plate'!B9",
        "B24": "='Data Plate'!B10",
        "B25": "='Data Plate'!B11",
        "B26": "='Data Plate'!B12",
        "B27": "='Data Plate'!B13",
        "B28": "='Data Plate'!B14",
        "B29": "='Data Plate'!B15",
        "B30": "='Data Plate'!B16",
    }
    for cell, expected_formula in dp_links.items():
        val = pc[cell].value
        check(f"Perf Calc {cell} links to Data Plate",
              val == expected_formula,
              f"expected {expected_formula!r}, got {val!r}")

    # ================================================================
    # 3. Computed Constants still reference local B19-B30
    # ================================================================
    # Sigma (B39) should reference B34 (h) and NOT 'Data Plate'
    sigma_formula = pc["B39"].value
    check("B39 (sigma) references B34",
          isinstance(sigma_formula, str) and "B34" in sigma_formula,
          f"got {sigma_formula!r}")
    check("B39 (sigma) does NOT ref Data Plate",
          isinstance(sigma_formula, str) and "Data Plate" not in sigma_formula,
          f"got {sigma_formula!r}")

    # Phi (B41) should reference B39 and B30
    phi_formula = pc["B41"].value
    check("B41 (phi) references B39 and B30",
          isinstance(phi_formula, str) and "B39" in phi_formula and "B30" in phi_formula,
          f"got {phi_formula!r}")

    # X (B42) should reference B26
    x_formula = pc["B42"].value
    check("B42 (X) references B26",
          isinstance(x_formula, str) and "B26" in x_formula,
          f"got {x_formula!r}")

    # SDF (B44) should reference B28 and B27
    sdf_formula = pc["B44"].value
    check("B44 (SDF) references B28 and B27",
          isinstance(sdf_formula, str) and "B28" in sdf_formula and "B27" in sdf_formula,
          f"got {sdf_formula!r}")

    # CP (B47) should reference B40, B46, B23
    cp_formula = pc["B47"].value
    check("B47 (CP) references B23 (d)",
          isinstance(cp_formula, str) and "B23" in cp_formula,
          f"got {cp_formula!r}")

    # P (B45) should reference B36 and B21
    p_formula = pc["B45"].value
    check("B45 (P) references B36 and B21",
          isinstance(p_formula, str) and "B36" in p_formula and "B21" in p_formula,
          f"got {p_formula!r}")

    # ================================================================
    # 4. Performance table spot checks
    # ================================================================
    # Row 95 = KCAS 60.0
    check("Perf table A95 = 60.0",
          pc["A95"].value == 60.0,
          f"got {pc['A95'].value!r}")

    # KTAS formula references B39 (sigma) — may be $B$39
    ktas_95 = pc["B95"].value
    check("B95 (KTAS) references B39",
          isinstance(ktas_95, str) and "B$39" in ktas_95,
          f"got {ktas_95!r}")

    # Dp formula references B24 (CD0) and B19 (S) — may be $B$24
    dp_95 = pc["K95"].value
    check("K95 (Dp) references B24 and B19",
          isinstance(dp_95, str) and "B$24" in dp_95 and "B$19" in dp_95,
          f"got {dp_95!r}")

    # Di formula references B33 (W), B19 (S), B43 (A), B25 (e)
    di_95 = pc["L95"].value
    check("L95 (Di) references B33, B19, B43, B25",
          isinstance(di_95, str)
          and "B$33" in di_95 and "B$19" in di_95
          and "B$43" in di_95 and "B$25" in di_95,
          f"got {di_95!r}")

    # ================================================================
    # 5. V-speed formulas reference performance table range
    # ================================================================
    vy_formula = pc["B55"].value
    check("B55 (Vy) references table range",
          isinstance(vy_formula, str) and "N95" in vy_formula and "N255" in vy_formula,
          f"got {vy_formula!r}")

    vm_formula = pc["B59"].value
    check("B59 (VM) references table range",
          isinstance(vm_formula, str) and "A95" in vm_formula,
          f"got {vm_formula!r}")

    # ================================================================
    # 6. Section header updated
    # ================================================================
    header = pc["D18"].value
    check("Section header says 'linked from Data Plate'",
          isinstance(header, str) and "linked" in header.lower() and "Data Plate" in header,
          f"got {header!r}")

    # ================================================================
    # 7. Fill colors: data plate cells should be green (RESULT_FILL)
    # ================================================================
    result_fill_color = "C6EFCE"
    for cell_ref in ["B19", "B20", "B21", "B22", "B23", "B24", "B25",
                     "B26", "B27", "B28", "B29", "B30"]:
        fill = pc[cell_ref].fill
        check(f"Perf Calc {cell_ref} has green fill",
              fill.start_color and fill.start_color.rgb
              and result_fill_color in str(fill.start_color.rgb),
              f"got rgb={fill.start_color.rgb if fill.start_color else 'None'}")

    # Operational variables should still be yellow (INPUT_FILL)
    input_fill_color = "FFF2CC"
    for cell_ref in ["B33", "B34", "B35", "B36"]:
        fill = pc[cell_ref].fill
        check(f"Perf Calc {cell_ref} has yellow fill",
              fill.start_color and fill.start_color.rgb
              and input_fill_color in str(fill.start_color.rgb),
              f"got rgb={fill.start_color.rgb if fill.start_color else 'None'}")

    # ── Summary ──
    total = passed + failed
    print(f"\n{passed}/{total} checks passed.")
    if failed:
        print(f"{failed} FAILED.")
        sys.exit(1)
    else:
        print("All checks passed!")


if __name__ == "__main__":
    main()
